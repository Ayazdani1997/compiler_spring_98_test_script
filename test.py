import glob
import itertools
import os
import shutil
import subprocess
import zipfile
import signal
from enum import Enum
from pandas import *

from shutil import *
import openpyxl

base_dir = os.getcwd()
testcases_dir = os.path.join(base_dir, "final_tests")
codes_dir = os.path.join(base_dir, 'codes/Phase2')
excel_name = "Phase2_Grades"
testcase_mapper_filename = "phase2_testcases.csv"
project_dir = os.path.join(base_dir, "project_dir")
java_class_source_dir = "src"
grammar_source_dir = java_class_source_dir
compiler_runner = 'Toorla'
compiler_runner_java_file = compiler_runner + '.java'
FAILURE = "BUILD FAILURE"
testcase_extension = ".trl"
COMPILED = "Compiled"
excel_extension = ".xlsx"
output_extension = ".out"
separator = '_'
temp_directory = 'temp'
NUMOFSTUDENTS = 2
version_loss_rate = 0.1
default_timeout_in_secs = 10
RUNSNUM = "#run"
USAGE_OF_MAVEN = "usage_of_maven"
newer_version_sign = '$'
compressed_code_extension = ".zip"
items_to_copy = {
    'files': {
        'Toorla.g4': os.path.join(project_dir, grammar_source_dir),
        'ToorlaCompiler.java': os.path.join(project_dir, java_class_source_dir)
    },
    'directories': {
        'toorla': os.path.join(project_dir, java_class_source_dir)
    }
}

pom_filename = 'pom.xml'
pom_file_content = '''<?xml version="1.0" encoding="UTF-8"?>
<project xmlns="http://maven.apache.org/POM/4.0.0"
         xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
         xsi:schemaLocation="http://maven.apache.org/POM/4.0.0 http://maven.apache.org/xsd/maven-4.0.0.xsd">
    <modelVersion>4.0.0</modelVersion>

    <groupId>toorla</groupId>
    <artifactId>student_project</artifactId>
    <version>1.0-SNAPSHOT</version>
    <properties>
        <maven.compiler.source>1.8</maven.compiler.source>
        <maven.compiler.target>1.8</maven.compiler.target>
    </properties>

    <build>
        <sourceDirectory>
            ''' + java_class_source_dir + '''
        </sourceDirectory>
        <plugins>
            <plugin>
                <groupId>org.antlr</groupId>
                <artifactId>antlr4-maven-plugin</artifactId>
                <version>4.7.2</version>
                <executions>
                    <execution>
                        <id>antlr</id>
                        <goals>
                            <goal>antlr4</goal>
                        </goals>
                        <configuration>
                            <sourceDirectory>''' + grammar_source_dir + '''</sourceDirectory>
                        </configuration>
                    </execution>
                </executions>
            </plugin>
        </plugins>
    </build>
    <dependencies>
        <dependency>
            <groupId>org.antlr</groupId>
            <artifactId>antlr4</artifactId>
            <version>4.7.2</version>
        </dependency>
    </dependencies>



</project> '''

max_run_num = 5
FULL_TESTCASE_GRADE = 1


class Grade(Enum):
    ERROR = -FULL_TESTCASE_GRADE
    FAULT = 0
    OK = FULL_TESTCASE_GRADE


class TimeOutException(Exception):
    def __str__(self):
        return "TIMEOUT OCCURRED!!!"


def create_new_student(worksheet, sid_list):
    row = ([0 for i in range(worksheet.max_column)])
    for i in range(len(sid_list)):
        row[i] = sid_list[i]
    usage_of_maven_index = find_col_index_in_sheet(worksheet, USAGE_OF_MAVEN)
    compiled_index = find_col_index_in_sheet(worksheet, COMPILED)
    row[compiled_index - 1] = "No"
    row[usage_of_maven_index - 1] = "No"
    return row


def find_col_index_in_sheet(worksheet, col_name):
    if col_name is None:
        return -1
    for i in range(1, worksheet.max_column + 1):
        if worksheet.cell(1, i).value == col_name:
            return i
    return -1


def find_sid_index_in_sheet(worksheet, sid):
    for i in range(1, worksheet.max_row + 1):
        if worksheet.cell(i, 1).value == sid:
            return i
    return -1


def grade_students(worksheet, test_case_name, grade, sid_list, version=0):
    col = find_col_index_in_sheet(worksheet, test_case_name)
    if col == -1:
        raise Exception("Test case does not exist!!")
    row = find_sid_index_in_sheet(worksheet, sid_list[0])
    if grade == grade.OK:
        if worksheet.cell(row, col).value == 0:
            worksheet.cell(row, col).value = FULL_TESTCASE_GRADE - version_loss_rate * version


def save_result(worksheet, grade, sid_list, test_case_name=None, version=0):
    if test_case_name is not None:
        test_case_name = test_case_name.split(testcase_extension)[0]
    row = find_sid_index_in_sheet(worksheet, sid_list[0])
    col = find_col_index_in_sheet(worksheet, COMPILED)
    if grade != Grade.ERROR:
        worksheet.cell(row, col).value = "Yes"
    else:
        return
    grade_students(worksheet, test_case_name, grade, sid_list, version)


def build_compiler():
    prev_path = os.getcwd()
    os.chdir(project_dir)
    if pom_filename not in os.listdir(os.getcwd()):
        os.chdir(prev_path)
        raise Exception("the directory you are in does not contain a maven project")
    print("############## building project ##############")
    compile_command = "mvn compile -Dmaven.compiler.source=1.8 -Dmaven.compiler.target=1.8"
    p = subprocess.Popen(compile_command, shell=True,
                         stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    os.chdir(prev_path)
    p.wait()
    out, err = p.communicate()
    if out is not None and FAILURE in str(out):
        print("############## code has compile error ##############")
        return False
    print("############## building done successfully ##############\n")
    return True


def evaluate(testcase_root, testcase_name, output, print_message=True):
    pure_output = output.replace("\r\n", "").replace("\n", "").replace("\t", "").replace(" ", "")
    try:
        with open(os.path.join(testcase_root, testcase_name.split(testcase_extension)[0] + output_extension),
                  "r") as text_file:  # HYPO: file exists
            expected = text_file.read()
            pure_expected = expected.replace("\r\n", "") \
                .replace("\n", "").replace("\t", "").replace(" ", "")

        text_file.close()
    except OSError as e:
        expected = output
    if pure_expected != pure_output:
        if print_message:
            print("TEST CASE", testcase_name, "FAILED!!!!!:\nEXPECTED:\n", expected, "\nOUTPUT:\n", output + "\n\n")
        return False
    else:
        if print_message:
            print("TEST CASE", testcase_name, "PASSED!\n\n")
        return True


def find_junk_files_in_dir_tree(directory):
    dir_tree = list(os.walk(directory))
    junk_files = []
    junk_dirs = [entity[0] for entity in dir_tree if os.path.basename(entity[0]) == "__MACOSX"]
    return junk_files, junk_dirs


def decompress(directory, filename):
    prev_path = os.getcwd()
    os.chdir(directory)
    if not os.path.isfile(os.path.join(filename)):
        os.chdir(prev_path)
        raise Exception('this file name does not exist on code repo')
    os.mkdir(temp_directory)
    copyfile(os.path.join(directory, filename), os.path.join(temp_directory, filename))
    zip_ref = zipfile.ZipFile(os.path.join(directory, filename), 'r')
    zip_ref.extractall(os.path.join(directory, temp_directory))
    zip_ref.close()
    os.remove(os.path.join(temp_directory, filename))
    os.chdir(prev_path)
    complete_path = os.path.join(directory, temp_directory)
    junk_files, junk_dirs = find_junk_files_in_dir_tree(complete_path)
    for _file in junk_files:
        os.remove(_file)
    for _dir in junk_dirs:
        shutil.rmtree(_dir)
    return complete_path


def copy_items_to_dest(cur_dir, items_to_copy):
    copied_items = set({})
    for (root, dirs, files) in os.walk(cur_dir):
        directory_key = os.path.basename(root)
        if directory_key in items_to_copy['directories'] and not ('dir', directory_key) in copied_items:
            shutil.copytree(root, os.path.join(items_to_copy['directories'].get(directory_key), directory_key))
            copied_items.add(('dir', directory_key))
        for file in files:
            if file in items_to_copy['files'] and not ('file', file) in copied_items:
                copyfile(os.path.join(root, file), os.path.join(items_to_copy['files'].get(file), file))
                copied_items.add(('file', file))


def handle_run_timeout(signum, frame):
    raise TimeOutException()


def partial_compile_test(testcase_root, testcase_name):
    prev_path = os.getcwd()
    os.chdir(project_dir)
    runner = compiler_runner
    run_command = "mvn -q exec:java -Dexec.mainClass=" + runner + " -Dexec.args=" + os.path.join(
        testcase_root, testcase_name)
    p = subprocess.Popen(run_command, shell=True,
                         stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    os.chdir(prev_path)
    signal.signal(signal.SIGALRM, handle_run_timeout)
    signal.alarm(default_timeout_in_secs)
    try:
        p.wait()
        signal.alarm(0)
    except TimeOutException as timeOutException:
        p.kill()
        print(timeOutException)
        raise TimeOutException()
    out, err = p.communicate()
    pure_output = str(out)[2: len(str(out)) - 1]. \
        replace("\\r\\n", "\r\n").replace("\\n", "\n").replace("\\t", "\t")
    pure_stderr = str(err)[2: len(str(err)) - 1]
    return pure_output, pure_stderr


def run_test(testcase_dir, testcase_name):
    print("############## running code with test case", testcase_name,
          "is started  ##############\n")
    output, stderr = partial_compile_test(testcase_dir, testcase_name)
    return output, stderr


def test_group_project(worksheet, sids, version):
    for (testcase_root, testcase_dirs, testcase_files) in os.walk(testcases_dir):
        for testcase_name in testcase_files:
            if testcase_name.endswith(testcase_extension):
                grade = Grade.FAULT
                try:
                    output, stderr = run_test(testcase_root, testcase_name)
                    if evaluate(testcase_root, testcase_name, output):
                        grade = Grade.OK
                    save_result(worksheet, grade, sids, testcase_name, version)
                except TimeOutException:
                    save_result(worksheet, grade, sids, testcase_name, version)


def get_sids(code_name):
    sid = code_name.split(compressed_code_extension)[0]
    return sid.split(separator)


def save_new_code_to(project_dir, student_code_dir, new_code_name):
    archive_address = shutil.make_archive(new_code_name, 'zip', project_dir)
    copyfile(archive_address, os.path.join(student_code_dir, new_code_name + compressed_code_extension))
    os.remove(archive_address)


def get_num_of_runs_for_std(sids, worksheet):
    row = find_sid_index_in_sheet(worksheet, sids[0])
    col = find_col_index_in_sheet(worksheet, RUNSNUM)
    if row == -1:
        worksheet.append(create_new_student(worksheet, sids))
        row = worksheet.max_row
    version = worksheet.cell(row, col).value
    return version


def extract_code(code_dir, compressed_code_name):
    return decompress(code_dir, compressed_code_name)


def create_antlr_maven_project_in(project_dir):
    shutil.rmtree(project_dir, ignore_errors=True)
    os.mkdir(project_dir)
    os.mkdir(os.path.join(project_dir, java_class_source_dir))
    try:
        os.mkdir(os.path.join(project_dir, grammar_source_dir))
    except OSError:
        pass
    pom_file = open(os.path.join(project_dir, pom_filename), 'w+')
    pom_file.write(pom_file_content)
    pom_file.close()


def extract_project_from_source(code_dir, code_name, version):
    maven_project_detected = True
    _version = version
    if _version >= max_run_num:
        raise Exception("max number of runs exceeded")
    new_code_name = ('' if _version is 0 else (newer_version_sign + str(_version) + '_')) + code_name
    if not new_code_name in os.listdir(code_dir):
        raise Exception("version " + str(_version + 1) + " of this code does not exist on repo")
    decompressed_code_location = extract_code(code_dir, new_code_name)
    project_pom_file_addresses = glob.glob(
        os.path.join(os.path.join(decompressed_code_location, '**'), pom_filename), recursive=True)
    if len(project_pom_file_addresses) != 0:
        pom_file_dir_address = os.path.dirname(project_pom_file_addresses[0])
        shutil.rmtree(project_dir, ignore_errors=True)
        shutil.copytree(pom_file_dir_address, project_dir)
    else:
        maven_project_detected = False
        create_antlr_maven_project_in(project_dir)
        copy(compiler_runner_java_file, os.path.join(project_dir, java_class_source_dir))
        copy_items_to_dest(decompressed_code_location, items_to_copy)
    rmtree(os.path.join(code_dir, decompressed_code_location))
    return maven_project_detected


def prepare_project(code_dir, code_name, version, copy_from_source):
    if not copy_from_source:
        no_project_found = False
        if os.path.basename(project_dir) in os.listdir(base_dir):
            if not pom_filename in os.listdir(project_dir):
                no_project_found = True
                shutil.rmtree(project_dir, ignore_errors=True)
        else:
            no_project_found = True
        if no_project_found:
            print("no project found to run, exiting")
            exit(1)
        return True
    else:
        return extract_project_from_source(code_dir, code_name, version)


def do_test_scenario(worksheet, code_dir, code_name, version, copy_from_source=True):
    sids = get_sids(code_name)
    std_row = find_sid_index_in_sheet(worksheet, sids[0])
    run_col = find_col_index_in_sheet(worksheet, RUNSNUM)
    if std_row == -1:
        worksheet.append(create_new_student(worksheet, sids))
        std_row = find_sid_index_in_sheet(worksheet, sids[0])
    maven_project_detected = prepare_project(code_dir, code_name, version, copy_from_source)
    _version = version
    if not copy_from_source:
        _version = get_num_of_runs_for_std(sids, worksheet)
    else:
        if version == 0:
            worksheet.cell(std_row, find_col_index_in_sheet(worksheet,
                                                            USAGE_OF_MAVEN)).value = "Yes" if maven_project_detected else "No"

    print("--------------------------------- group ", ','.join(sids), "-------------------------------------")
    compiled = build_compiler()
    if not compiled:
        save_result(worksheet, Grade.ERROR, sids)
    else:
        test_group_project(worksheet, sids, _version)
    first_version_pure_name = separator.join(sids)
    new_pure_code_name = (first_version_pure_name if _version == 0
                          else newer_version_sign + str(_version) + "_" + first_version_pure_name)
    if not copy_from_source:
        save_new_code_to(project_dir, code_dir, new_pure_code_name)
    shutil.rmtree(project_dir, ignore_errors=True)
    print("---------------------------------- end of running tests of group", ', '.join(sids), "------------------")
    worksheet.cell(std_row, run_col).value = worksheet.cell(std_row, run_col).value + 1


def test_all(worksheet, version=None):
    recent_students = set({})
    code_dir = codes_dir
    for (code_root, codeDirs, code_files) in os.walk(code_dir):
        for code_name in code_files:
            if code_name.endswith(compressed_code_extension) and not code_name.startswith(newer_version_sign):
                sids = get_sids(code_name)
                if len(set(sids) & recent_students) is 0:
                    recent_students |= set(sids)
                    try:
                        _version = version
                        if version is None:
                            _version = get_num_of_runs_for_std(sids, worksheet)
                        do_test_scenario(worksheet, code_root, code_name, _version)
                    except Exception as exception:
                        print(exception)


def get_test_addr_by_id(test_id):
    try:
        tests_csv_file = pandas.read_csv(testcase_mapper_filename)
        testcase_position = list(tests_csv_file['id']).index(test_id)
        testcase_root = tests_csv_file['testcase_dir'][testcase_position]
        testcase_name = tests_csv_file['testcase_name'][testcase_position]
        return testcase_root, testcase_name
    except Exception:
        print('no test with such id')
        exit(1)


def try_test(code_dir, code_name, test_id, version, copy_from_source=True):
    sids = get_sids(code_name)
    if copy_from_source:
        print("########## trying test %d on version %d of code"
              " for group %s ##########" % (test_id, version + 1, '_'.join(sids)))
    else:
        print("########## trying test %d on living code #############" % test_id)
    testcase_root, testcase_name = get_test_addr_by_id(test_id)
    prepare_project(code_dir, code_name, version, copy_from_source)
    compiled = build_compiler()
    if compiled:
        try:
            output, stderr = run_test(testcase_root, testcase_name)
            evaluate(testcase_root, testcase_name, output)
        except TimeOutException:
            pass


def find_sid_code_name(sids):
    for (code_root, codeDirs, code_files) in os.walk(codes_dir):
        for code_name in code_files:
            filename_list = [separator.join(list(element)) + compressed_code_extension for element in
                             list(itertools.permutations(sids))]
            for filename in filename_list:
                if code_name == filename:
                    return code_root, code_name
    return None, None


def create_default_row():
    row1 = [('sid' + str(i + 1)) for i in range(NUMOFSTUDENTS)]
    row1.extend([COMPILED, USAGE_OF_MAVEN, RUNSNUM])
    row2 = ["810199XXX" for i in range(NUMOFSTUDENTS)]
    row2.extend(["Yes", "Yes", 0])
    for (testcase_root, testcase_dirs, testcase_files) in os.walk(testcases_dir):
        for testcase_name in testcase_files:
            if testcase_name.endswith(testcase_extension):
                pure_testcase_name = testcase_name.split(testcase_extension)[0]
                row1.append(pure_testcase_name)
                row2.append(0)
    return [row1, row2]


def create_excel(excel_name):
    try:
        workbook = openpyxl.load_workbook(excel_name)
        worksheet = workbook.active
    except:
        number_tests()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = excel_name.split(excel_extension)[0]
        header_rows = create_default_row()
        worksheet.append(header_rows[0])
        worksheet.append(header_rows[1])
    return worksheet, workbook


def is_there_any_change_in_testcases_dir(testcase_mapper_columns):
    if testcase_mapper_filename in os.listdir(base_dir):
        cur_testcases_name = set({})
        tests_csv_file = pandas.read_csv(testcase_mapper_filename)
        if set(testcase_mapper_columns) == set(tests_csv_file.columns):
            for (root, dirs, files) in os.walk(testcases_dir):
                for test_file in files:
                    if test_file.endswith(testcase_extension):
                        cur_testcases_name.add(test_file)
            csv_tests_name = set(
                {testcase_name for testcase_name in tests_csv_file['testcase_name']})
            if cur_testcases_name == csv_tests_name:
                return False
    return True


def number_tests():
    test_id = 0
    testcase_mapper = {'id': [], 'testcase_dir': [], 'testcase_name': []}
    change_detected = is_there_any_change_in_testcases_dir({'id', 'testcase_dir', 'testcase_name'})
    if change_detected:
        for (test_dir, test_dirs, test_files) in os.walk(testcases_dir):
            for test_file in test_files:
                if test_file.endswith(testcase_extension):
                    test_id += 1
                    testcase_name = test_file.split(testcase_extension)[0]
                    new_testcase_name = str(test_id) + '_' + testcase_name
                    os.rename(os.path.join(test_dir, test_file),
                              os.path.join(test_dir, new_testcase_name + testcase_extension))
                    output_name = testcase_name + output_extension
                    new_output_name = new_testcase_name + output_extension
                    os.rename(os.path.join(test_dir, output_name), os.path.join(test_dir, new_output_name))
                    testcase_mapper.get('id').append(test_id)
                    testcase_mapper.get('testcase_dir').append(test_dir)
                    testcase_mapper.get('testcase_name').append(new_testcase_name + testcase_extension)
        dataframe = DataFrame(testcase_mapper, columns=['id', 'testcase_dir', 'testcase_name'])
        dataframe.to_csv(testcase_mapper_filename, index=None, header=True)


def check_for_testcases_format():
    test_id = 0
    try:
        for (test_root, test_dirs, test_files) in os.walk(testcases_dir):
            for test_file in test_files:
                if test_file.endswith(testcase_extension):
                    test_id += 1
                if test_file not in os.listdir(test_root):
                    raise Exception(
                        'test file with name ' + test_file + '\'s output is missing')
    except Exception as exception:
        print(exception)
        exit(1)


def list_tests(prefix=None):
    testcases = pandas.read_csv(testcase_mapper_filename, header=0)
    print(DataFrame(
        [testcase for testcase in testcases['testcase_name'] if prefix is None or testcase.startswith(prefix)],
        columns=['testcase_name']))


def remove_duplicate_codes(codes_dir, compressed_code_extension='.zip', newer_version_sign='$'):
    recent_students = set({})
    code_dir = codes_dir
    for (code_root, codeDirs, code_files) in os.walk(code_dir):
        for code_name in code_files:
            if code_name.endswith(compressed_code_extension) and not code_name.startswith(newer_version_sign):
                sids = get_sids(code_name)
                if len(set(sids) & recent_students) is 0:
                    recent_students |= set(sids)
                else:
                    os.remove(os.path.join(code_root, code_name))


def list_groups():
    group_id = 0
    code_dir = codes_dir
    for (code_root, codeDirs, code_files) in os.walk(code_dir):
        for code_name in code_files:
            if code_name.endswith(compressed_code_extension) and not code_name.startswith(newer_version_sign):
                group_id += 1
                sids = get_sids(code_name)
                print('group', group_id, ':', ' , '.join(sids))


def check_for_prerequisites():
    if os.path.basename(testcases_dir) not in os.listdir(base_dir) or not os.path.isdir(testcases_dir):
        raise Exception('There are no test cases in your system')
    elif not os.path.isdir(codes_dir):
        raise Exception('There are not codes in your system to test')
    elif compiler_runner_java_file not in os.listdir(base_dir) or not os.path.isfile(compiler_runner_java_file):
        raise Exception('There is no compiler entry point ( java runner class ) to copy to non maven projects')


def try_all_codes(test_id, version):
    for (code_root, code_dirs, code_files) in os.walk(codes_dir):
        for code_file in code_files:
            if code_file.endswith(compressed_code_extension) and not code_file.startswith(newer_version_sign):
                try:
                    try_test(code_root, code_file, test_id, version)
                    print("########## end of trial ############\n\n")
                except Exception as version_not_exists:
                    print(version_not_exists)


def parse_try_command(command):
    print("enter test id :")
    test_id = int(input())
    version = 0
    try:
        if not '-all-codes' in command:
            copy_from_source = False
            group_code_dir, code_name = "", ""
            if '-noCopyFromSource' not in command:
                copy_from_source = True
                print("enter sids : ")
                sids = input().split()
                if len(sids) <= 0:
                    print("you must enter at least one sid")
                    return
                group_code_dir, code_name = find_sid_code_name(sids)
                if group_code_dir is None or code_name is None:
                    print('no code with such student-ids')
                    return
                print("enter version number : ")
                version_input = input()
                version = int(version_input)
            try_test(group_code_dir, code_name, test_id, (version - 1 if version > 0 else 0), copy_from_source)
        else:
            print("enter version number : ")
            version_input = input()
            version = int(version_input)
            try_all_codes(test_id, (version - 1 if version > 0 else 0))
    except Exception as version_not_exists:
        print(version_not_exists)


def parse_test_command(command, worksheet, excel_name):
    if not '-all-codes' in command:
        print("enter sids : ")
        sids = input().split()
        if len(sids) <= 0:
            print("you must enter at least one sid")
            return
        group_code_dir, code_name = find_sid_code_name(sids)
        if group_code_dir is None or code_name is None:
            print('no code with such student-ids')
            return
        copy_from_source = True
        if '-noCopyFromSource' in command:
            copy_from_source = False
        sids = get_sids(code_name)
        try:
            version = get_num_of_runs_for_std(sids, worksheet)
            do_test_scenario(worksheet, group_code_dir, code_name, version, copy_from_source)
        except Exception as exception:
            print(exception)
    else:
        test_all(worksheet)
    workbook.save(excel_name)


if __name__ == "__main__":
    check_for_prerequisites()
    check_for_testcases_format()
    remove_duplicate_codes(codes_dir)
    excel_file_name = excel_name + excel_extension
    worksheet, workbook = create_excel(excel_file_name)
    workbook.save(excel_file_name)
    help = "commands: \n\n" \
           "test: tests code with specified sids located in codes directory which is hard coded . if it comes with " \
           + "-noCopyFromSource, it tests the living code on " \
           + "project dir\n" \
           + "\toption all-codes : tests all codes located in code_dir dir path," \
           + "try: it just runs a test for a group after getting test id and student ids, if it comes " \
             "with " \
             "-noCopyFromSource it runs the living code on project_dir\n" \
             "\toption all-codes : runs a test for all group" \
           + "\nhelp( h ) : prints this manual\n" \
           + "exit: termination of cli\n" \
           + "list_tests: lists all tests available in test case directory with name\n" \
           + "list_groups: list all groups who sent you code\n\n"
    while True:
        print('>>>>>', end=" ")
        command = str(input())
        if command == 'exit':
            print("######### bye , see you soon! #########")
            break
        elif command.startswith('try'):
            parse_try_command(command)
        elif command.startswith('test'):
            parse_test_command(command, worksheet, excel_file_name)
        elif command == 'help' or command == 'h':
            print(help)
        elif command.startswith('list_tests'):
            list_tests()
        elif command.startswith('list_groups'):
            list_groups()
        else:
            print("unknown command")
            print(help)
